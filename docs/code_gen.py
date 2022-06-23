import openpyxl
from openpyxl.cell import Cell

import json

api_doc_file = "KIS Developers API명세서_220616.xlsx"
wb = openpyxl.load_workbook(api_doc_file)
print(wb.sheetnames)

sheet_names = ['OAuth인증', '국내주식주문', '국내주식시세', '국내주식실시간',
               '국내선물옵션주문', '국내선물옵션시세', '해외주식주문', '해외주식현재가']
sheet = wb[sheet_names[2]]


def extract_api_list_from_sheet():
    api_row_nums = []
    count = 0
    for row in sheet.iter_rows(values_only=False):
        if row[0].value == "API명":
            # print(len(row), row[0].row)
            api_row_nums.append([row[1].value, row[0].row, 0])
            if count > 0:
                api_row_nums[count - 1][2] = row[0].row
            count += 1

    api_row_nums[-1][2] = sheet.max_row
    return api_row_nums


def extract_api_parameters(name):
    _row_index = 0
    _header = {}
    _found = {"API명": False, "Request": False}
    for row in sheet.iter_rows(values_only=False):

        # print(row[1].value, name)
        if not _found["API명"] and row[0].value == "API명" and row[1].value.strip() == name.strip():
            _row_index = row[0].row
            print("/// " + row[1].value)
            _found["API명"] = True
        elif not _found["API명"]:
            continue

        if not _found["Request"] and _found["API명"]:
            if row[0].value == "Request":
                _found["Request"] = True
                continue
            else:
                continue

        if _found["Request"] and row[0].value and row[0].value.strip() in ["Header", "Body"]:
            continue

        if _found["Request"] and row[0].value == "tr_id" and row[5].value == "[실전/모의투자]":
            _header[f"{row[0].value}"] = \
                f"{sheet[row[0].row + 1][5].value}".split(":")[0].strip()

            continue

        if _found["Request"]:
            _header[f"{row[0].value}"] = f"{row[5].value}"

        if _found["Request"] and row[0].value == "Response":
            break

    return _row_index, _header


def get_api_basic_info(row_index):
    # api_name = f"{sheet[row_index][0].value}"
    api_info = f"{sheet[row_index + 1][1].value}"
    api_params_raw = api_info.split("\n")
    api_params = {}
    for data in api_params_raw:
        api_params[data.split(":")[0].strip()] = data.split(":")[1].strip()

    return api_params


def generate_code(api_summary, header):
    print("pub fn API_NAME_XXX(&self) -> KisResult<serde_json::Value> {")
    print(f"    let url = \"{api_summary['URL']}\";")
    print(f'    let headers = [("tr_id", "{header["tr_id"]}")];')
    print(
        f'    let query = [("fid_cond_mrkt_div_code", "J"), ("fid_input_iscd", ticker)];')
    print(
        f'    let req = self.make_request(url, RequestType::{api_summary["Method"]}, &headers, &query)?;')
    print('')
    print(f'    self.send_request(req)')
    print("}")


if __name__ == '__main__':
    # All api info in a sheet
    # api_list = extract_api_list_from_sheet()
    # for item in api_list:
    #     # print('"' + item[0] + '"' + ",")
    #     print("- [ ] "+item[0])
    # exit()
    apis_name_list = [
        "주식현재가 시세[v1_국내주식-008]",
        "주식현재가 체결[v1_국내주식-009]",
        "주식현재가 일자별[v1_국내주식-010]",
        "주식현재가 호가 예상체결[v1_국내주식-011]",
        "주식현재가 투자자[v1_국내주식-012]",
        "주식현재가 회원사[v1_국내주식-013]",
        "ELW현재가 시세[v1_국내주식-014]",
        "국내주식기간별시세(일/주/월/년)[v1_국내주식-016]",
        "국내주식업종기간별시세(일/주/월/년)[v1_국내주식-021]",
    ]
    for api_name in apis_name_list:
        begin_row_num, header = extract_api_parameters(api_name)
        summary = get_api_basic_info(begin_row_num)
        generate_code(summary, header)
